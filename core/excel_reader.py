from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import urlparse

EXCEL_PATH = Path(__file__).parent.parent / "data" / "Prensa Deportiva.xlsx"

# Tu mapa de categorías original
_CAT_MAP = {
    "REAL MADRID":                  "Real Madrid Masculino",
    "REAL MADRID FEMENINO":         "Real Madrid Femenino",
    "LALIGA":                       "LaLiga",
    "SELECCIÓN ESPAÑOLA":           "Selección Española Masculina",
    "SELECCIÓN ESPAÑOLA FEMENINO":  "Selección Española Femenina",
    "LIGA FUTVE":                   "Liga FUTVE",
    "VINOTINTO":                    "Vinotinto Masculina",
    "VENEZUELA FEMENINO":           "Vinotinto Femenina",
    "VENEZUELA":                    "General",
    # ── Categorías del Mundial 2026 ──────────────────────────────────────────
    "MUNDIAL GENERAL":              "Mundial General",
    "SELECCIONES LATINAS":          "Selecciones Latinas",
    "FASE DE GRUPOS":               "Fase de Grupos",
    "PORTUGAL Y EUROPEOS":          "Portugal y Europeos",
}

def _friendly_name(url: str) -> str:
    domain = urlparse(url).netloc.lower().removeprefix("www.")
    parts = domain.split(".")
    if len(parts) >= 2:
        return parts[-2].capitalize()
    return parts[0].capitalize()

def load_sources() -> dict[str, list[dict]]:
    sources = {}

    # 1. Cargar el Excel principal de VG
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        current_cat = "General"
        for row in ws.iter_rows(values_only=True):
            cell = row[0]
            if not cell: continue
            value = str(cell).strip()
            if value.endswith(":") and value == value.upper():
                title = value.rstrip(":").strip()
                current_cat = _CAT_MAP.get(title, title)
                continue
            if value.startswith("http"):
                sources.setdefault(current_cat, []).append(
                    {"nombre": _friendly_name(value), "url": value}
                )
        wb.close()

    # 2. Cargar el Excel del Mundial de forma dinámica (ignorar mayúsculas/minúsculas)
    data_dir = Path(__file__).parent.parent / "data"
    mundial_path = None
    if data_dir.exists():
        for file in data_dir.glob("*.xlsx"):
            fname = file.name.lower()
            # Si el archivo contiene "mundial" o "enlaces"
            if "mundial" in fname and "lista" in fname:
                mundial_path = file
                break
                
        if not mundial_path:
            # Plan B: buscar cualquier archivo que tenga "mundial" y "enlaces"
            for file in data_dir.glob("*.*"):
                fname = file.name.lower()
                if "mundial" in fname and ("lista" in fname or "enlaces" in fname):
                    mundial_path = file
                    break

    if mundial_path and mundial_path.exists():
        try:
            wb_m = load_workbook(mundial_path, data_only=True)
            ws_m = wb_m.active
            for row in ws_m.iter_rows(values_only=True):
                cell = row[0]
                if not cell: continue
                value = str(cell).strip()
                if value.startswith("http"):
                    # Asignar todo a la categoría "Mundial Global"
                    sources.setdefault("Mundial Global", []).append(
                        {"nombre": _friendly_name(value), "url": value}
                    )
            wb_m.close()
        except Exception as e:
            print(f"Error cargando Excel Mundial: {e}")

    return sources