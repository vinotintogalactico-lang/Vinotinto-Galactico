from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import urlparse

EXCEL_PATH = Path(__file__).parent.parent / "data" / "Prensa Deportiva.xlsx"

_CAT_MAP = {
    "REAL MADRID":                  "Real Madrid Masculino",
    "REAL MADRID FEMENINO":         "Real Madrid Femenino",
    "LALIGA":                       "LaLiga",
    "SELECCIÓN ESPAÑOLA":           "Selección Española Masculina",
    "SELECCIÓN ESPAÑOLA FEMENINO":  "Selección Española Femenina",
    "LIGA FUTVE":                   "Liga FUTVE",
    "VINOTINTO":                    "Vinotinto Masculina",
    "VENEZUELA FEMENINO":           "Vinotinto Femenina",
    "VENEZUELA":                    None,
}

CATEGORY_ORDER = [
    "Real Madrid Masculino", "Real Madrid Femenino", "LaLiga",
    "Selección Española Masculina", "Selección Española Femenina",
    "Vinotinto Masculina", "Vinotinto Femenina", "Liga FUTVE",
]

def _friendly_name(url: str) -> str:
    domain = urlparse(url).netloc.lower().removeprefix("www.")
    mapping = {
        "realmadrid.com": "Real Madrid Oficial", "as.com": "AS",
        "marca.com": "Marca", "mundodeportivo.com": "Mundo Deportivo",
        "defensacentral.com": "Defensa Central", "madridistareal.com": "Madridista Real",
        "sport.es": "Sport", "estadiodeportivo.com": "Estadio Deportivo",
        "laliga.com": "LaLiga Oficial", "meridiano.net": "Meridiano",
        "liderendeportes.com": "Líder en Deportes", "lavinotinto.com": "La Vinotinto",
        "balonazos.com": "Balonazos", "fvf.com.ve": "FVF Oficial",
    }
    for key, name in mapping.items():
        if key in domain:
            return name
    return domain.split(".")[0].capitalize()

def load_sources() -> dict[str, list[dict]]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontró el Excel en: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    sources: dict[str, list[dict]] = {}
    current_cat: str | None = None

    for row in ws.iter_rows(values_only=True):
        cell = row[0]
        if cell is None:
            continue
        value = str(cell).strip()
        if not value:
            continue
        if value.endswith(":") and value == value.upper():
            current_cat = _CAT_MAP.get(value.rstrip(":").strip())
            continue
        if value.startswith("http") and current_cat is not None:
            sources.setdefault(current_cat, []).append(
                {"nombre": _friendly_name(value), "url": value}
            )

    wb.close()

    ordered: dict[str, list[dict]] = {}
    for cat in CATEGORY_ORDER:
        if cat in sources:
            ordered[cat] = sources[cat]
    for cat in sources:
        if cat not in ordered:
            ordered[cat] = sources[cat]
    return ordered