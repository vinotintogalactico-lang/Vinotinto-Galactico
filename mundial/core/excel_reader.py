from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import urlparse

EXCEL_PATH = Path(__file__).parent.parent / "data" / "Prensa_Mundial2026.xlsx"

_CAT_MAP = {
    "TV / CANAL":      "📺 TV / Canal",
    "PRENSA ESCRITA":  "📰 Prensa Escrita",
    "DIGITAL":         "💻 Digital",
}


_CUSTOM_NAMES = {
    "foxsports.com.mx":       "Fox Sports MX",
    "ole.com.ar":             "Olé",
    "elgrafico.com.ar":       "El Gráfico",
    "elpais.com.uy":          "El País UY",
    "record.com.mx":          "Record MX",
    "eluniversal.com.mx":     "El Universal MX",
    "studiofutbol.com.ec":    "Studio Fútbol",
    "espndeportes.espn.com":  "ESPN Deportes",
    "espn.com":               "ESPN Deportes",
    "noticiasrcn.com":        "RCN Noticias",
    "caracoltv.com":          "Caracol TV",
    "mitelefe.com":           "Telefe",
    "directvgo.com":          "DirecTV",
    "elcanaldelfutbol.com":   "Canal del Fútbol",
    "tvazteca.com":           "TV Azteca",
    "telemundodeportes.com":  "Telemundo",
    "chilevision.cl":         "Chilevision",
    "alairelibre.cl":         "Al Aire Libre",
    "latercera.com":          "La Tercera",
    "eltiempo.com":           "El Tiempo CO",
    "clarosports.com":        "Claro Sports",
    "liderendeportes.com":    "Líder en Deportes",
    "winsports.co":           "Win Sports",
    "futbolred.com":          "Futbol Red",
    "mediotiempo.com":        "Mediotiempo",
    "redgol.cl":              "RedGol",
    "libero.pe":              "Líbero",
    "diez.hn":                "Diez HN",
    "13.cl":                  "Canal 13 CL",
    "latina.pe":              "Latina PE",
    "rtve.es":                "RTVE",
    "tycsports.com":          "TyC Sports",
    "tudn.com":               "TUDN",
    "meridiano.net":          "Meridiano",
    "marca.com":              "Marca",
    "as.com":                 "AS",
    "mundodeportivo.com":     "Mundo Deportivo",
    "sport.es":               "Sport ES",
    "relevo.com":             "Relevo",
    "infobae.com":            "Infobae",
    "depor.com":              "Depor",
    "televen.com":            "Televen",
}


def _friendly_name(url: str) -> str:
    domain = urlparse(url).netloc.lower().removeprefix("www.")
    for key, name in _CUSTOM_NAMES.items():
        if key in domain:
            return name
    # Fallback: tomar la parte significativa del dominio
    parts = domain.split(".")
    # Quitar TLDs multi-parte (.com.ar → buscar antes de .com)
    skip = {"com", "net", "org", "co", "mx", "ar", "uy", "pe", "cl", "hn", "ec", "es"}
    for part in parts:
        if part not in skip and len(part) > 2:
            return part.capitalize()
    return parts[0].capitalize()


def load_sources() -> dict[str, list[dict]]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontró el Excel en: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    sources: dict[str, list[dict]] = {}
    current_cat = "💻 Digital"  # fallback

    for row in ws.iter_rows(values_only=True):
        cell = row[0]
        if not cell:
            continue
        value = str(cell).strip()

        # Detectar encabezado de sección (termina en ":")
        if value.endswith(":"):
            title = value.rstrip(":").strip().upper()
            current_cat = _CAT_MAP.get(title, title)
            continue

        # Detectar URL
        if value.startswith("http"):
            sources.setdefault(current_cat, []).append(
                {"nombre": _friendly_name(value), "url": value}
            )

    wb.close()
    return sources
