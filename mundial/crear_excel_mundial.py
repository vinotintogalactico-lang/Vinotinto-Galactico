"""
Script de un solo uso: genera data/Prensa_Mundial2026.xlsx
con las 3 secciones (TV/CANAL, PRENSA ESCRITA, DIGITAL)
"""
from openpyxl import Workbook
from pathlib import Path

data = {
    "TV / CANAL:": [
        "https://www.telemundodeportes.com/copa-mundial-fifa-2026",
        "https://www.tudn.com/mundial-2026",
        "https://www.tvazteca.com/aztecadeportes/futbol",
        "https://televen.com/elnoticiero/copa-mundial-de-la-fifa-2026/",
        "https://noticias.caracoltv.com/golcaracol/mundial-2026",
        "https://www.noticiasrcn.com/deportes",
        "https://www.tycsports.com/mundial",
        "https://mitelefe.com/deportes/",
        "https://www.directvgo.com/deportes",
        "https://elcanaldelfutbol.com/",
        "https://www.chilevision.cl/deportes",
        "https://www.13.cl/deportes",
        "https://www.latina.pe/deportes",
        "https://www.rtve.es/play/deportes/",
        "https://www.foxsports.com.mx/futbol/",
        "https://www.clarosports.com/futbol/mundial/",
    ],
    "PRENSA ESCRITA:": [
        "https://meridiano.net/futbol/futbol-internacional/",
        "https://www.record.com.mx/futbol",
        "https://www.ole.com.ar/mundial/mundial-2026",
        "https://www.elgrafico.com.ar/",
        "https://redgol.cl/mundial/",
        "https://depor.com/futbol-internacional/",
        "https://libero.pe/futbol-internacional/",
        "https://www.elpais.com.uy/ovacion/futbol",
        "https://www.diez.hn/futbolinternacional/",
        "https://www.marca.com/futbol/mundial/",
        "https://as.com/futbol/mundial/",
        "https://www.mundodeportivo.com/futbol/mundial",
        "https://www.sport.es/es/mundial/",
        "https://www.relevo.com/futbol/mundial-futbol/",
        "https://www.eltiempo.com/deportes/futbol-internacional",
        "https://www.latercera.com/canal/el-deportivo/",
        "https://www.eluniversal.com.mx/deportes/",
    ],
    "DIGITAL:": [
        "https://www.liderendeportes.com/noticias/futbol/mundial-2026/",
        "https://www.winsports.co/futbol-internacional/",
        "https://www.futbolred.com/futbol-internacional",
        "https://www.mediotiempo.com/futbol",
        "https://www.alairelibre.cl/noticias/deportes/futbol/mundial/",
        "https://studiofutbol.com.ec/category/internacional/",
        "https://espndeportes.espn.com/futbol/copa-mundial/",
        "https://www.infobae.com/deportes/",
    ],
}

wb = Workbook()
ws = wb.active
ws.title = "Enlaces"
ws["A1"] = "Prensa Mundial 2026 - Lista de Fuentes"

row = 2
for section, links in data.items():
    ws.cell(row=row, column=1, value=section)
    row += 1
    for link in links:
        ws.cell(row=row, column=1, value=link)
        row += 1
    row += 1  # línea en blanco entre secciones

out = Path(__file__).parent / "data" / "Prensa_Mundial2026.xlsx"
out.parent.mkdir(exist_ok=True)
wb.save(out)
print(f"✅ Excel generado: {out}")
print(f"   Total links: {sum(len(v) for v in data.values())}")
